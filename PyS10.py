# -*- coding: utf-8 -*-
"""
Created on Wed Dec 13 15:54:36 2023

@author: Diego
"""
import openpyxl
from tqdm import tqdm
import tkinter as tk
from tkinter import filedialog

def run_script():
    input_file = input_entry.get()
    output_file = output_entry.get()  # Get the output file name from the entry widget
    
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    last_row = ws.max_row

    headers = ['Item',
       'Partida',
       'Rendimiento',
       'Und_Rend',
       'Tipo_Recurso',
       'eliminar',
       'Recurso',
       'eliminar',
       'eliminar',
       'Und_Recurso',
       'Cuadrilla',
       'Cantidad',
       'Precio_Unit',
       'Precio_Parcial']

    new_sheet.append(headers)

    recursos_partidas = []
    item = 0

    listado_recurso = ('mano de obra', 'materiales', 'equipos')

    #iter_rows uses index 1
    for index, row in tqdm(enumerate(ws.iter_rows(min_row=1, max_col=9, values_only= True)),total=last_row):

        #rows as a list use index 0    
        if row[0] and "partida" in str(row[0]).lower():

            #cell uses index 1
            item = item + 1
            titulo = ws.cell(index + 1, 4).value
            rendimiento = ws.cell(index + 3, 3).value
            unidad_rendimiento = ws.cell(index + 3, 2).value
            
            i = index + 2
            
            while not("partida" in str(ws.cell(i, 1).value).lower()):
                
                if  i == last_row:
                    
                    break
                
                if str(ws.cell(i, 3).value).lower() in listado_recurso:
                    
                    tipo_recurso = str(ws.cell(i, 3).value)
                
                elif ws.cell(i, 2).value and not("rendimiento" in str(ws.cell(i,1).value).lower()) and not("código" in str(ws.cell(i,1).value).lower()):
                    
                    valores = [
                        item,
                        titulo,
                        rendimiento,
                        unidad_rendimiento
                        ]
                    valores.append(str(tipo_recurso).upper())
                    row_values = list(ws.values)[i - 1]
                    valores.extend(row_values)
                    recursos_partidas.append(valores)
                i = i + 1
            
            for recursos in recursos_partidas:
                
                new_sheet.append(recursos)
                
            recursos_partidas = []
            valores = []

    new_wb.save(output_file)

    root.destroy()

# Create Tkinter window
root = tk.Tk()
root.title("Excel Script GUI")

# Input file path
tk.Label(root, text="Input File Path:").pack()
input_entry = tk.Entry(root, width=50)
input_entry.pack()
tk.Button(root, text="Browse", command=lambda: input_entry.insert(tk.END, filedialog.askopenfilename())).pack()

# Output file name
tk.Label(root, text="Output File Name:").pack()
output_entry = tk.Entry(root, width=50)
output_entry.pack()
tk.Button(root, text="Browse", command=lambda: output_entry.insert(tk.END, filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    filetypes=[("Excel files", "*xlsx"), ("All files, ","*.*")]
    ))).pack()

# Run button
tk.Button(root, text="Run Script", command=run_script).pack()

root.mainloop()
